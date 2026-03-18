"""
测试执行模块

@input Config, TestCase, MemorySegment (配置和类型)
@input hardware_detector.quick_hardware_check (硬件预检测)
@output TestExecutor类, TestResult类
@pos 核心模块，负责DSS测试执行、结果收集、断点续测、硬件连接检测

一旦我被更新务必更新我的开头注释以及所属文件夹的 README.md
"""

import os
import json
import csv
import shutil
import tempfile
import subprocess
import time
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Any
from dataclasses import dataclass

from .config import Config, TestCase, MemorySegment
from .exceptions import (
    TestError,
    DSSNotFoundError,
    TestExecutionError,
)
from .logger import get_logger, LogContext
from .hardware_detector import quick_hardware_check

logger = get_logger(__name__)


@dataclass
class TestResult:
    """测试结果"""
    case_name: str
    status: str
    dat_dir: Optional[Path] = None
    error: Optional[str] = None


class TestExecutor:
    """
    测试执行器
    
    用法：
        executor = TestExecutor(config)
        results = executor.run_all()
        for r in results:
            print(f"{r.case_name}: {r.status}")
    """
    
    def __init__(self, config: Config, run_timestamp: Optional[str] = None):
        self.config = config
        self.dss_exe = config.paths.ccs_dss
        self.ccxml = config.paths.ccxml
        self.device = config.test.device
        self.cpu = config.test.cpu
        self.timeout = config.test.test_timeout
        self.batch_size = config.test.test_batch_size
        self.result_addr = config.test.result_addr
        self.success_val = config.test.success_val
        self.error_val = config.test.error_val
        
        self.template_path = Path(__file__).parent.parent / "templates" / "dss_test.js.tmpl"
        # 如果提供了时间戳则使用，否则生成新的（用于断点续测时保持时间戳一致）
        self.run_timestamp = run_timestamp if run_timestamp else datetime.now().strftime("%Y-%m-%d-%H-%M")
        
        # 回调函数
        self.on_case_started: Optional[callable] = None
        self.on_case_finished: Optional[callable] = None
        self.on_hardware_error: Optional[callable] = None  # 硬件连接错误回调
    
    def validate(self):
        """验证测试环境"""
        if not self.dss_exe.exists():
            raise DSSNotFoundError(str(self.dss_exe))
        if not self.template_path.exists():
            raise TestError(
                f"DSS 模板文件不存在: {self.template_path}",
                error_code=4007,
                details={"template_path": str(self.template_path)}
            )
    
    def generate_test_config(self, output_path: Optional[Path] = None) -> Path:
        """
        从工作空间生成测试配置
        
        Args:
            output_path: 输出文件路径，默认为 full_regr.json
        
        Returns:
            生成的配置文件路径
        """
        with LogContext(logger, "生成测试配置"):
            # 优先使用配置中已有的用例
            if self.config.cases:
                logger.info(f"使用配置中的 {len(self.config.cases)} 个用例")
                return self._generate_test_config_from_cases(output_path)
            
            # 否则从工作空间搜索 .out 文件
            return self._generate_test_config_from_workspace(output_path)
    
    def _generate_test_config_from_cases(self, output_path: Optional[Path] = None) -> Path:
        """
        从配置中的用例生成测试配置
        
        Args:
            output_path: 输出文件路径
        
        Returns:
            生成的配置文件路径
        """
        cases = []
        for case in self.config.cases:
            # 确保 dat_dir 有有效值
            if case.dat_dir:
                dat_dir = case.dat_dir
            else:
                dat_dir = f"5_result_dat/{self.run_timestamp}/{case.name}"
            
            cases.append({
                "name": case.name,
                "out": str(case.out).replace("\\", "/"),
                "dat_dir": dat_dir,
                "segments": [
                    {"name": s.name, "addr": s.addr, "len": s.len, "width": s.width}
                    for s in (case.segments if case.segments else self.config.memory_segments)
                ]
            })
        
        test_config = {
            "ccxml": str(self.ccxml).replace("\\", "/"),
            "device": self.device,
            "cpu": self.cpu,
            "timeout": self.timeout,
            "result_addr": self.result_addr,
            "success_val": self.success_val,
            "error_val": self.error_val,
            "cases": cases
        }
        
        if output_path is None:
            output_path = Path("full_regr.json")
        
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(test_config, f, ensure_ascii=False, indent=2)
        
        logger.info(f"测试配置已保存: {output_path}")
        return output_path
    
    def _generate_test_config_from_workspace(self, output_path: Optional[Path] = None) -> Path:
        """
        从工作空间搜索 .out 文件生成测试配置
        
        Args:
            output_path: 输出文件路径
        
        Returns:
            生成的配置文件路径
        """
        workspace = self.config.paths.ccs_workspace
        
        out_files = sorted(workspace.rglob("*.out"))
        
        if not out_files:
            logger.warning(f"工作空间中没有找到 .out 文件: {workspace}")
            return None
        
        logger.info(f"找到 {len(out_files)} 个 .out 文件")
        
        cases = []
        for out_file in out_files:
            case_name = out_file.stem
            dat_dir = f"5_result_dat/{self.run_timestamp}/{case_name}"
            
            cases.append({
                "name": case_name,
                "out": str(out_file.resolve()).replace("\\", "/"),
                "dat_dir": dat_dir,
                "segments": [
                    {"name": s.name, "addr": s.addr, "len": s.len, "width": s.width}
                    for s in self.config.memory_segments
                ]
            })
        
        test_config = {
            "ccxml": str(self.ccxml).replace("\\", "/"),
            "device": self.device,
            "cpu": self.cpu,
            "timeout": self.timeout,
            "result_addr": self.result_addr,
            "success_val": self.success_val,
            "error_val": self.error_val,
            "cases": cases
        }
        
        if output_path is None:
            output_path = Path("full_regr.json")
        
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(test_config, f, ensure_ascii=False, indent=2)
        
        logger.info(f"测试配置已保存: {output_path}")
        return output_path
    
    def run_all(self, test_config_path: Optional[Path] = None, 
                start_batch: int = 0,
                resume_from_last: bool = False) -> List[TestResult]:
        """
        执行所有测试
        
        Args:
            test_config_path: 测试配置文件路径，如果为 None 则自动生成
            start_batch: 从第几个批次开始执行（0-based），用于断点续测
            resume_from_last: 是否从上次中断的批次继续（会覆盖 start_batch）
        
        Returns:
            测试结果列表
        """
        with LogContext(logger, "测试执行"):
            self.validate()
            
            if test_config_path is None:
                test_config_path = self.generate_test_config()
            
            if test_config_path is None or not test_config_path.exists():
                logger.error("没有测试配置文件，跳过测试执行")
                return []
            
            with open(test_config_path, "r", encoding="utf-8") as f:
                test_config = json.load(f)
            
            cases = test_config.get("cases", [])
            if not cases:
                logger.warning("测试配置中没有测试用例")
                return []
            
            logger.info(f"共 {len(cases)} 个测试用例")
            
            # 检查是否需要断点续测
            # 注意：如果用户已经指定了 start_batch（>0），则使用用户指定的值
            # 只有 start_batch=0 且 resume_from_last=True 时才自动检测
            if resume_from_last and start_batch == 0:
                detected_batch = self._find_last_completed_batch(cases)
                if detected_batch > 0:
                    start_batch = detected_batch
                    logger.info(f"自动检测到已完成的批次，从第 {start_batch + 1} 批继续执行")
            elif start_batch > 0:
                logger.info(f"使用用户指定的起始批次，从第 {start_batch + 1} 批继续执行")
            
            self._create_output_dirs(cases)
            
            log_dir = Path("6_result_dat_logs") / self.run_timestamp
            log_dir.mkdir(parents=True, exist_ok=True)
            
            batches = self._split_into_batches(cases, self.batch_size)
            logger.info(f"分成 {len(batches)} 批执行 (每批 {self.batch_size} 个)")
            
            # 如果指定了开始批次，跳过前面的
            if start_batch > 0:
                logger.info(f"跳过前 {start_batch} 个批次")
                batches = batches[start_batch:]
            
            console_log = log_dir / "console_all.log"
            all_results: List[TestResult] = []
            
            # 用例开始时间字典
            case_start_times: Dict[str, float] = {}
            
            with open(console_log, "a", encoding="utf-8") as f_out:
                for i, batch in enumerate(batches, start_batch + 1):
                    logger.info(f"执行第 {i}/{start_batch + len(batches)} 批 ({len(batch)} 个用例)")
                    
                    # 通知批次中的用例开始，并记录开始时间
                    logger.info(f"通知批次开始，共 {len(batch)} 个用例")
                    for case in batch:
                        logger.info(f"通知用例开始: {case['name']}")
                        case_start_times[case["name"]] = time.time()
                        if self.on_case_started:
                            try:
                                self.on_case_started(case["name"])
                                logger.info(f"已通知用例开始: {case['name']}")
                            except Exception as e:
                                logger.error(f"通知用例开始失败: {case['name']}, 错误: {e}")
                    
                    # 批次前检测硬件连接
                    if not self._check_hardware_connection():
                        logger.error("=" * 60)
                        logger.error(f"第 {i} 批执行前检测到硬件连接断开！")
                        logger.error("=" * 60)
                        logger.error("请检查以下项目:")
                        logger.error("  1. XDS100 调试器是否正确连接到电脑")
                        logger.error("  2. 目标板是否上电")
                        logger.error("  3. FTDI 驱动是否安装正确")
                        logger.error("  4. USB 线缆是否正常")
                        logger.error("=" * 60)
                        logger.error(f"已执行 {i-1} 个批次，可以从第 {i} 批继续执行")
                        logger.error(f"命令: python run.py test --start-batch {i-1}")
                        
                        # 通知当前批次用例完成（连接断开状态）
                        for case in batch:
                            if self.on_case_finished:
                                self.on_case_finished(case["name"], "ConnectionLost", 0.0)
                        
                        # 记录后续批次为未执行（跳过当前批次，从下一个开始）
                        current_batch_index = batches.index(batch)
                        if current_batch_index + 1 < len(batches):
                            self._mark_remaining_batches_as_skipped(batches[current_batch_index + 1:])
                        
                        # 触发硬件错误回调
                        if self.on_hardware_error:
                            try:
                                error_msg = f"第 {i} 批执行前检测到硬件连接断开"
                                self.on_hardware_error(i, error_msg)
                            except Exception as e:
                                logger.error(f"触发硬件错误回调失败: {e}")
                        
                        break
                    
                    logger.info(f"硬件连接正常，开始执行第 {i} 批")
                    batch_start_time = time.time()
                    dss_success = self._run_batch(batch, test_config, log_dir, f_out)
                    batch_end_time = time.time()
                    logger.info(f"第 {i} 批 DSS 执行完成，成功: {dss_success}")
                    
                    # 批次执行后再次检测硬件连接
                    # 如果执行前正常但执行后失败，说明执行过程中硬件断开
                    logger.info(f"开始第 {i} 批执行后硬件连接检测...")
                    try:
                        hardware_connected_after = self._check_hardware_connection()
                        logger.info(f"第 {i} 批执行后硬件连接状态: {'正常' if hardware_connected_after else '断开'}")
                    except Exception as e:
                        logger.error(f"第 {i} 批执行后硬件连接检测异常: {e}")
                        hardware_connected_after = False
                    
                    # 立即收集该批次的结果并通知完成
                    logger.info(f"收集批次结果并通知完成")
                    batch_results = self._collect_batch_results(batch)
                    logger.info(f"批次结果: {len(batch_results)} 个")
                    
                    # 判断批次是否成功：
                    # 1. DSS执行成功
                    # 2. 有结果文件
                    # 3. 执行后硬件连接正常（如果执行前正常的话）
                    batch_actually_success = dss_success and len(batch_results) == len(batch) and hardware_connected_after
                    
                    # 处理有结果的用例
                    result_case_names = {r.case_name for r in batch_results}
                    for result in batch_results:
                        all_results.append(result)
                        # 计算耗时
                        start_time = case_start_times.get(result.case_name, batch_start_time)
                        duration = batch_end_time - start_time
                        logger.info(f"Case finished: {result.case_name} = {result.status}, duration: {duration:.2f}s")
                        if self.on_case_finished:
                            try:
                                self.on_case_finished(result.case_name, result.status, duration)
                                logger.info(f"已通知用例完成: {result.case_name}")
                            except Exception as e:
                                logger.error(f"通知用例完成失败: {result.case_name}, 错误: {e}")
                    
                    # 处理没有结果的用例（执行失败或未生成结果文件）
                    for case in batch:
                        if case["name"] not in result_case_names:
                            # 判断是否是硬件断开导致的失败
                            if not hardware_connected_after:
                                logger.error(f"用例未生成结果文件（硬件断开）: {case['name']}")
                                status = "ConnectionLost"
                            else:
                                logger.error(f"用例未生成结果文件，标记为失败: {case['name']}")
                                status = "Failed"
                            duration = batch_end_time - case_start_times.get(case["name"], batch_start_time)
                            if self.on_case_finished:
                                try:
                                    self.on_case_finished(case["name"], status, duration)
                                    logger.info(f"已通知用例{status}: {case['name']}")
                                except Exception as e:
                                    logger.error(f"通知用例失败失败: {case['name']}, 错误: {e}")
                    
                    if not batch_actually_success:
                        logger.error(f"第 {i} 批执行失败（DSS成功: {dss_success}, 结果数: {len(batch_results)}/{len(batch)}, 硬件连接: {'正常' if hardware_connected_after else '断开'}），停止后续批次")
                        
                        # 检查是否是硬件连接问题导致的执行失败
                        # 情况1：执行后硬件连接断开
                        # 情况2：DSS执行失败（包括超时、返回错误码等）
                        # 情况3：DSS返回成功但没有结果文件
                        if not hardware_connected_after or not dss_success or (dss_success and len(batch_results) == 0):
                            logger.error(f"检测到硬件连接中断或执行失败，触发硬件错误回调")
                            if self.on_hardware_error:
                                try:
                                    if not hardware_connected_after:
                                        error_msg = f"第 {i} 批执行过程中检测到硬件连接断开"
                                    elif not dss_success:
                                        error_msg = f"第 {i} 批DSS执行失败（超时或错误），可能是硬件连接问题"
                                    else:
                                        error_msg = f"第 {i} 批执行失败，可能是硬件连接中断（未生成结果文件）"
                                    self.on_hardware_error(i, error_msg)
                                except Exception as e:
                                    logger.error(f"触发硬件错误回调失败: {e}")
                            else:
                                logger.error(f"on_hardware_error callback is None!")
                        
                        # 记录后续批次为未执行
                        self._mark_remaining_batches_as_skipped(batches[batches.index(batch)+1:])
                        break
            
            # 收集剩余未执行用例的结果（如果有）
            executed_case_names = {r.case_name for r in all_results}
            for case in cases:
                if case["name"] not in executed_case_names:
                    # 检查该用例是否有结果文件
                    case_result = self._read_case_result(case)
                    if case_result:
                        all_results.append(case_result)
            
            # 按用例名称排序结果，保持顺序一致
            case_name_to_result = {r.case_name: r for r in all_results}
            sorted_results = []
            for case in cases:
                if case["name"] in case_name_to_result:
                    sorted_results.append(case_name_to_result[case["name"]])
            
            self._generate_summary_report(cases, sorted_results)
            
            self._post_process_dat_files(cases)
            
            self._cleanup_old_logs()
            
            return sorted_results
    
    def _find_last_completed_batch(self, cases: List[Dict]) -> int:
        """
        查找最后一个完成的批次索引
        
        逻辑：
        - 如果批次中所有用例都有结果文件且状态为 Success，则认为该批次已完成
        - 如果批次中有用例没有结果文件，或状态不是 Success，则认为该批次未完成
        - 返回第一个未完成的批次索引，用于断点续测
        
        Args:
            cases: 所有用例列表
            
        Returns:
            第一个未完成的批次索引（0-based），如果全部完成则返回批次总数
        """
        batches = self._split_into_batches(cases, self.batch_size)
        
        for i, batch in enumerate(batches):
            # 检查该批次的所有用例是否都有成功的结果
            for case in batch:
                summary_file = Path(case["dat_dir"]) / "summary.csv"
                if not summary_file.exists():
                    # 该批次未完成（没有结果文件）
                    logger.info(f"断点续测检测: 批次 {i+1} 用例 {case['name']} 无结果文件，从该批次开始")
                    return i
                
                # 检查结果状态
                try:
                    with open(summary_file, "r", encoding="utf-8") as f:
                        content = f.read().strip()
                        if not content:
                            logger.info(f"断点续测检测: 批次 {i+1} 用例 {case['name']} 结果文件为空，从该批次开始")
                            return i
                        
                        # 解析 CSV 内容
                        parts = content.split(",")
                        if len(parts) >= 2:
                            status = parts[1]
                            # 只有 Success 状态才算完成
                            if status != "Success":
                                logger.info(f"断点续测检测: 批次 {i+1} 用例 {case['name']} 状态为 {status}，从该批次开始")
                                return i
                        else:
                            logger.info(f"断点续测检测: 批次 {i+1} 用例 {case['name']} 结果格式错误，从该批次开始")
                            return i
                except Exception as e:
                    logger.info(f"断点续测检测: 批次 {i+1} 用例 {case['name']} 读取结果失败: {e}，从该批次开始")
                    return i
            
            # 该批次所有用例都成功完成，继续检查下一批次
            logger.info(f"断点续测检测: 批次 {i+1} 已完成")
        
        # 所有批次都已完成
        logger.info(f"断点续测检测: 所有 {len(batches)} 个批次已完成")
        return len(batches)
    
    def _mark_remaining_batches_as_skipped(self, batches: List[List[Dict]]):
        """
        将剩余批次标记为跳过
        
        Args:
            batches: 剩余批次列表
        """
        for batch in batches:
            for case in batch:
                self._write_case_result(case, "Skipped")
    
    def _collect_batch_results(self, batch: List[Dict]) -> List[TestResult]:
        """
        收集批次中各用例的结果
        
        Args:
            batch: 批次中的用例列表
            
        Returns:
            测试结果列表
        """
        results = []
        logger.info(f"收集批次结果，批次包含 {len(batch)} 个用例")
        for case in batch:
            logger.info(f"收集用例结果: {case['name']}")
            result = self._read_case_result(case)
            if result:
                results.append(result)
                logger.info(f"成功收集用例结果: {case['name']} = {result.status}")
            else:
                logger.info(f"未能收集用例结果: {case['name']}")
        logger.info(f"批次结果收集完成，共 {len(results)} 个结果")
        return results
    
    def _read_case_result(self, case: Dict) -> Optional[TestResult]:
        """
        读取单个用例的结果
        
        Args:
            case: 用例配置
            
        Returns:
            测试结果，如果没有结果则返回 None
        """
        # 使用绝对路径
        dat_dir = Path(case["dat_dir"]).resolve()
        summary_file = dat_dir / "summary.csv"
        
        logger.info(f"尝试读取用例结果: {case['name']} -> {summary_file}")
        
        if not summary_file.exists():
            logger.info(f"结果文件不存在: {summary_file}")
            return None
        
        try:
            with open(summary_file, "r", encoding="utf-8") as f:
                content = f.read().strip()
                if not content:
                    logger.info(f"结果文件为空: {summary_file}")
                    return None
                
                # 解析 CSV 内容
                parts = content.split(",")
                if len(parts) >= 2:
                    case_name = parts[0]
                    status = parts[1]
                    logger.info(f"用例 {case_name} 结果: {status}")
                    return TestResult(
                        case_name=case_name,
                        status=status,
                        dat_dir=Path(case["dat_dir"])
                    )
        except Exception as e:
            logger.warning(f"读取用例结果失败 {case['name']}: {e}")
        
        return None
    
    def _create_output_dirs(self, cases: List[Dict]):
        """创建输出目录"""
        for case in cases:
            dat_dir = Path(case["dat_dir"])
            dat_dir.mkdir(parents=True, exist_ok=True)
    
    def _split_into_batches(self, cases: List[Dict], batch_size: int) -> List[List[Dict]]:
        """将用例分批"""
        batches = []
        for i in range(0, len(cases), batch_size):
            batches.append(cases[i:i + batch_size])
        return batches
    
    def _check_hardware_connection(self) -> bool:
        """
        快速检查硬件连接状态
        
        优化策略：
        1. 首先进行 USB 设备预检测（快速，< 1秒）
        2. 如果预检测通过，再进行 DSS 连接检测（验证实际可连接性）
        3. 如果预检测失败，直接返回失败，避免等待 DSS 超时
        
        Returns:
            连接是否正常
        """
        # 步骤1: USB 设备预检测（快速）
        logger.info("执行硬件预检测...")
        precheck_passed, precheck_msg = quick_hardware_check()
        
        if not precheck_passed:
            logger.warning(f"硬件预检测失败: {precheck_msg}")
            # 预检测失败，直接返回，避免调用耗时的 DSS 检测
            return False
        
        logger.info(f"硬件预检测通过: {precheck_msg}")
        
        # 步骤2: DSS 连接检测（验证实际可连接性）
        # 只有在预检测通过后才执行，此时硬件大概率存在，DSS 检测会很快
        try:
            # 创建临时连接检测脚本
            check_script = self._create_connection_check_script()
            
            with tempfile.NamedTemporaryFile(mode="w", suffix=".js", delete=False, encoding="utf-8") as f:
                f.write(check_script)
                js_file = f.name
            
            try:
                cmd = [str(self.dss_exe), js_file]
                logger.info(f"执行 DSS 硬件检测命令: {' '.join(cmd)}")
                result = subprocess.run(
                    cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    timeout=10,  # 快速检测，10秒超时
                    creationflags=subprocess.CREATE_NO_WINDOW
                )
                
                # 检查输出中是否包含成功标志
                output = result.stdout.decode('utf-8', errors='ignore')
                stderr_output = result.stderr.decode('utf-8', errors='ignore')
                
                logger.info(f"DSS 硬件检测返回码: {result.returncode}")
                logger.info(f"DSS 硬件检测输出: {output}")
                if stderr_output:
                    logger.info(f"DSS 硬件检测错误输出: {stderr_output}")
                
                # 检查连接状态：必须包含 CONNECTION_OK 且不包含 CONNECTION_FAILED
                is_connected = "CONNECTION_OK" in output and "CONNECTION_FAILED" not in output and result.returncode == 0
                
                if not is_connected:
                    logger.warning(f"DSS 硬件连接检测失败，返回码: {result.returncode}")
                    logger.warning(f"标准输出: {output[:500]}")
                    if stderr_output:
                        logger.warning(f"错误输出: {stderr_output[:500]}")
                
                return is_connected
                
            finally:
                try:
                    os.unlink(js_file)
                except:
                    pass
                    
        except Exception as e:
            logger.warning(f"DSS 硬件连接检测失败: {e}")
            import traceback
            logger.warning(f"检测异常详情: {traceback.format_exc()}")
            return False
    
    def _create_connection_check_script(self) -> str:
        """
        创建硬件连接检测脚本
        
        Returns:
            JavaScript 脚本内容
        """
        ccxml_path = str(self.ccxml).replace('\\', '/')
        device_name = self.device
        cpu_name = self.cpu
        return f'''
importPackage(Packages.com.ti.debug.engine.scripting);
importPackage(Packages.com.ti.ccstudio.scripting.environment);
importPackage(Packages.java.lang);

print("正在连接目标设备...");

var env = null;
var server = null;
var session = null;
var connected = false;

try {{
    // 创建调试会话
    env = ScriptingEnvironment.instance();
    server = env.getServer("DebugServer.1");
    
    // 先尝试停止任何现有的调试会话（清理残留状态）
    try {{
        server.stop();
        print("已停止现有调试会话");
    }} catch (e) {{
        // 忽略错误，可能没有现有会话
    }}
    
    server.setConfig("{ccxml_path}");

    // 打开会话
    session = server.openSession("{device_name}", "{cpu_name}");

    // 尝试连接
    print("尝试连接目标...");
    try {{
        session.target.connect();
        connected = session.target.isConnected();
    }} catch (e) {{
        print("CONNECTION_FAILED: 连接异常 - " + e.message);
        connected = false;
    }}

    // 检查连接状态
    if (connected) {{
        print("CONNECTION_OK: 硬件连接正常");
        try {{
            session.target.disconnect();
        }} catch (e) {{
            print("警告: 断开连接时出错 - " + e.message);
        }}
    }} else {{
        print("CONNECTION_FAILED: 硬件连接失败");
    }}
}} catch (e) {{
    print("CONNECTION_FAILED: 脚本异常 - " + e.message);
}} finally {{
    // 确保资源被释放
    if (session) {{
        try {{
            if (session.target.isConnected()) {{
                session.target.disconnect();
            }}
        }} catch (e) {{}}
    }}
    if (server) {{
        try {{
            server.stop();
        }} catch (e) {{}}
    }}
}}

print("检测完成");
'''

    def _run_batch(self, batch: List[Dict], global_config: Dict, log_dir: Path, log_file) -> bool:
        """
        执行一批测试用例
        
        Args:
            batch: 批次中的用例列表
            global_config: 全局配置
            log_dir: 日志目录
            log_file: 日志文件
            
        Returns:
            是否成功执行该批次
        """
        log_file_path = str((log_dir / f"DSS_Batch_{batch[0]['name']}.xml").resolve()).replace("\\", "/")
        
        config_json = {
            "global": {
                "ccxml": global_config["ccxml"],
                "device": global_config["device"],
                "cpu": global_config["cpu"]
            },
            "log_file": log_file_path,
            "cases": [
                {
                    "case_name": case["name"],
                    "out": case["out"],
                    "timeout": global_config["timeout"],
                    "result_addr": int(global_config["result_addr"], 16),
                    "success_val": int(global_config["success_val"], 16),
                    "error_val": int(global_config["error_val"], 16),
                    "dat_dir": str(Path(case["dat_dir"]).resolve()).replace("\\", "/"),
                    "segments": [
                        {
                            "name": s["name"],
                            "addr": int(s["addr"], 16),
                            "len": int(s["len"], 16),
                            "width": s["width"]
                        }
                        for s in case["segments"]
                    ]
                }
                for case in batch
            ]
        }
        
        js_script = self._generate_js_script(config_json)
        
        with tempfile.NamedTemporaryFile(mode="w", suffix=".js", delete=False, encoding="utf-8") as f:
            f.write(js_script)
            js_file = f.name
        
        try:
            cmd = [str(self.dss_exe), js_file]
            logger.info(f"执行 DSS 命令: {' '.join(cmd)}")
            result = subprocess.run(
                cmd,
                stdout=log_file,
                stderr=subprocess.STDOUT,
                timeout=len(batch) * self.timeout / 1000 + 60
            )
            
            # 检查 DSS 执行结果
            if result.returncode != 0:
                logger.warning(f"批次 DSS 执行返回非零码: {result.returncode}")
                # 检查是否是连接错误
                if self._check_connection_error_in_log(log_dir):
                    return False
            
            return True
                
        except subprocess.TimeoutExpired:
            logger.error(f"批次 DSS 执行超时")
            return False
        finally:
            try:
                os.unlink(js_file)
            except Exception as e:
                logger.warning(f"清理临时 JS 文件失败: {e}")
    
    def _check_connection_error_in_log(self, log_dir: Path) -> bool:
        """
        检查日志中的连接错误
        
        Args:
            log_dir: 日志目录
            
        Returns:
            是否检测到连接错误
        """
        console_log_path = log_dir / "console_all.log"
        
        if not console_log_path.exists():
            return False
            
        try:
            with open(console_log_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                # 检查各种连接错误标志
                connection_error_patterns = [
                    "Error connecting to the target",
                    "emulation failure",
                    "TARGET_CONNECT_FAILED",
                    "FTDI driver functions",
                    "no XDS100 is plugged in"
                ]
                for pattern in connection_error_patterns:
                    if pattern in content:
                        logger.error("=" * 60)
                        logger.error("检测到硬件连接问题！")
                        logger.error("=" * 60)
                        logger.error("请检查以下项目:")
                        logger.error("  1. XDS100 调试器是否正确连接到电脑")
                        logger.error("  2. 目标板是否上电")
                        logger.error("  3. FTDI 驱动是否安装正确")
                        logger.error("  4. USB 线缆是否正常")
                        logger.error("=" * 60)
                        return True
        except Exception as e:
            logger.warning(f"检查连接错误时出错: {e}")
        
        return False
    
    def _write_case_result(self, case: Dict, status: str):
        """写入单个用例的结果"""
        try:
            summary_file = Path(case["dat_dir"]) / "summary.csv"
            summary_file.parent.mkdir(parents=True, exist_ok=True)
            with open(summary_file, "w", encoding="utf-8") as f:
                f.write(f"{case['name']},{status}\n")
        except Exception as e:
            logger.warning(f"写入用例结果失败: {e}")
    
    def _generate_js_script(self, config_json: Dict) -> str:
        """生成 DSS JavaScript 脚本"""
        with open(self.template_path, "r", encoding="utf-8") as f:
            template = f.read()
        
        config_str = json.dumps(config_json, indent=2)
        return template.replace("${CONFIG_JSON}", config_str)
    
    def _collect_results(self, cases: List[Dict]) -> List[TestResult]:
        """收集测试结果"""
        results = []
        
        if not cases:
            return results
        
        # 从第一个用例的 dat_dir 获取时间戳目录
        first_dat_dir = Path(cases[0]["dat_dir"])
        time_dir = first_dat_dir.parent
        summary_file = time_dir / "summary.csv"
        
        logger.debug(f"查找汇总文件: {summary_file}")
        
        # 检查时间戳目录是否存在
        if not time_dir.exists():
            logger.warning(f"时间戳目录不存在: {time_dir}")
            # 为所有用例返回无结果
            for case in cases:
                results.append(TestResult(
                    case_name=case["name"],
                    status="NoResult",
                    error=f"时间戳目录不存在: {time_dir}"
                ))
            return results
        
        # 读取汇总文件
        if summary_file.exists():
            try:
                with open(summary_file, "r", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        if len(row) >= 2:
                            case_name = row[0]
                            status = row[1]
                            # 找到对应的用例配置以获取 dat_dir
                            dat_dir = None
                            for case in cases:
                                if case["name"] == case_name:
                                    dat_dir = Path(case["dat_dir"])
                                    break
                            if dat_dir is None:
                                dat_dir = first_dat_dir  # 默认使用第一个用例的目录
                            
                            results.append(TestResult(
                                case_name=case_name,
                                status=status,
                                dat_dir=dat_dir
                            ))
                            logger.info(f"用例 {case_name} 结果: {status}")
                        else:
                            logger.warning(f"汇总文件行格式不正确: {row}")
            except Exception as e:
                logger.error(f"读取汇总文件失败: {e}")
                # 为所有用例返回无结果
                for case in cases:
                    results.append(TestResult(
                        case_name=case["name"],
                        status="NoResult",
                        error=f"读取汇总文件失败: {e}"
                    ))
        else:
            logger.warning(f"汇总文件不存在: {summary_file}")
            # 为所有用例返回无结果
            for case in cases:
                results.append(TestResult(
                    case_name=case["name"],
                    status="NoResult",
                    error="汇总文件不存在"
                ))
        
        return results
    
    def _generate_summary_report(self, cases: List[Dict], results: List[TestResult]):
        """生成汇总报告"""
        if not results:
            return
        
        # 找到第一个有有效 dat_dir 的结果
        time_dir = None
        for result in results:
            if result.dat_dir is not None:
                time_dir = result.dat_dir.parent
                break
        
        # 如果没有有效的 dat_dir，使用默认路径
        if time_dir is None:
            time_dir = Path("5_result_dat") / self.run_timestamp
            time_dir.mkdir(parents=True, exist_ok=True)
        
        summary_csv = time_dir / "summary.csv"
        
        # 生成完整的 CSV 汇总文件（包含所有用例）
        # 注意：DSS 脚本只生成批次内的汇总，这里生成总的汇总
        with open(summary_csv, "w", encoding="utf-8") as f:
            for result in results:
                f.write(f"{result.case_name},{result.status}\n")
        
        success_count = sum(1 for r in results if r.status == "Success")
        total_count = len(results)
        
        logger.info("=" * 60)
        logger.info(f"测试汇总: {success_count}/{total_count} 成功")
        logger.info(f"报告文件: {summary_csv}")
        logger.info("=" * 60)
        
        self._generate_excel_report(time_dir, results)
    
    def _generate_excel_report(self, time_dir: Path, results: List[TestResult]):
        """生成 Excel 报告"""
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("pandas/openpyxl 未安装，跳过 Excel 生成")
            return
        
        summary_csv = time_dir / "summary.csv"
        if not summary_csv.exists():
            return
        
        df = pd.read_csv(summary_csv, names=["TestCase", "Status"])
        df.insert(0, "序号", range(1, len(df) + 1))
        df = df[["序号", "TestCase", "Status"]]
        
        xlsx = summary_csv.with_suffix(".xlsx")
        df.to_excel(xlsx, index=False)
        
        wb = load_workbook(xlsx)
        ws = wb.active
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
            cell = row[0]
            if cell.value == "Success":
                cell.fill = green
            elif cell.value in ["Error", "Timeout", "Exception"]:
                cell.fill = red
        
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_len + 2
        
        wb.save(xlsx)
        logger.info(f"Excel 报告已生成: {xlsx}")
    
    def _post_process_dat_files(self, cases: List[Dict]):
        """后处理 .dat 文件（删除文件头）"""
        for case in cases:
            memory_dir = Path(case["dat_dir"]) / "memory"
            if memory_dir.exists():
                for dat_file in memory_dir.glob("*.dat"):
                    self._remove_first_line(dat_file)
    
    def _remove_first_line(self, file_path: Path):
        """删除文件第一行"""
        try:
            with open(file_path, "r", encoding="utf-8") as src:
                lines = src.readlines()
            with open(file_path, "w", encoding="utf-8") as dst:
                dst.writelines(lines[1:])
        except Exception as e:
            logger.debug(f"处理文件失败 {file_path}: {e}")
    
    def _cleanup_old_logs(self, days: int = 30):
        """清理旧日志"""
        import time
        cutoff = time.time() - days * 86400
        log_root = Path("6_result_dat_logs")
        
        if not log_root.exists():
            return
        
        for d in log_root.iterdir():
            if d.is_dir() and d.stat().st_mtime < cutoff:
                shutil.rmtree(d)
                logger.info(f"已清理旧日志: {d}")
