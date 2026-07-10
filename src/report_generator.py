"""
测试报告生成模块

负责生成 CSV 和 Excel 格式的测试报告

一旦我被更新务必更新我的开头注释以及所属文件夹的 README.md
"""

import csv
from pathlib import Path
from typing import List, Dict, Optional
from dataclasses import dataclass

from .logger import get_logger

logger = get_logger(__name__)


@dataclass
class TestResult:
    """测试结果"""
    case_name: str
    status: str
    dat_dir: Optional[Path] = None
    error: Optional[str] = None


def generate_summary_report(run_timestamp: str, cases: List[Dict], results: List[TestResult]) -> None:
    """
    生成汇总报告（CSV + Excel）

    Args:
        run_timestamp: 运行时间戳
        cases: 用例配置列表
        results: 测试结果列表
    """
    if not results:
        return

    # 找到第一个有有效 dat_dir 的结果
    time_dir = None
    for result in results:
        if result.dat_dir is not None:
            time_dir = result.dat_dir.parent
            break

    # 如果没有有效的 dat_dir，使用默认路径
    if time_dir is None:
        time_dir = Path("5_result_dat") / run_timestamp
        time_dir.mkdir(parents=True, exist_ok=True)

    summary_csv = time_dir / "summary.csv"

    # 生成完整的 CSV 汇总文件（包含所有用例）
    # 注意：DSS 脚本只生成批次内的汇总，这里生成总的汇总
    with open(summary_csv, "w", encoding="utf-8") as f:
        for result in results:
            f.write(f"{result.case_name},{result.status}\n")

    success_count = sum(1 for r in results if r.status == "Success")
    total_count = len(results)

    logger.info("=" * 60)
    logger.info(f"测试汇总: {success_count}/{total_count} 成功")
    logger.info(f"报告文件: {summary_csv}")
    logger.info("=" * 60)

    _generate_excel_report(time_dir, results)


def _generate_excel_report(time_dir: Path, results: List[TestResult]) -> None:
    """
    生成 Excel 报告

    Args:
        time_dir: 时间戳目录
        results: 测试结果列表
    """
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("pandas/openpyxl 未安装，跳过 Excel 生成")
        return

    summary_csv = time_dir / "summary.csv"
    if not summary_csv.exists():
        return

    df = pd.read_csv(summary_csv, names=["TestCase", "Status"])
    df.insert(0, "序号", range(1, len(df) + 1))
    df = df[["序号", "TestCase", "Status"]]

    xlsx = summary_csv.with_suffix(".xlsx")
    df.to_excel(xlsx, index=False)

    wb = load_workbook(xlsx)
    ws = wb.active
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        cell = row[0]
        if cell.value == "Success":
            cell.fill = green
        elif cell.value in ["Error", "Timeout", "Exception"]:
            cell.fill = red

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(xlsx)
    logger.info(f"Excel 报告已生成: {xlsx}")
